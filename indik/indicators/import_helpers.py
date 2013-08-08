from openpyxl import load_workbook
from .models import IndicatorDescriptor, IndicatorData
from unidecode import unidecode
from xlrd import open_workbook


def fix_string(name):
		"""
		"""
		return unidecode(name).replace(" ", "_")


class XLSXLoader(object):



	def load_xlsx_file(self, filename):
		wb = open_workbook(filename)
		#print wb
		info = self.load_info(wb)
		meta = self.load_meta(wb)
		klasses = self.load_klasses(wb)
		data = self.load_data(wb)

		return info, meta, klasses, data

		

	def load_info(self, wb):
		sheet = wb.sheet_by_name('info')
		name = sheet.cell(1,1).value
		code = sheet.cell(0,1).value
		return { 'code':code, 'name':name }

	def load_meta(self, wb):
		out = []
		sheet = wb.sheet_by_name('meta')

		for nrow in range(sheet.nrows):
			field = {}
			row = sheet.row(nrow)
			field_name = fix_string(row[0].value)
			field_type = row[1].value

			field = { 'name' : field_name, 'field_type':field_type}
			out.append(field)
		
		return out


	def load_data(self, wb):
		sheet = wb.sheet_by_name('dati')		
		header  =sheet.row(0)
		fields = []
		data = []
		for cell in header:
			field_name = fix_string(cell.value)
			fields.append(field_name)
		
		for nrow in range(1, sheet.nrows):
			item = dict(zip(fields, [x.value for x in sheet.row(nrow)]))
			data.append(item)
		return data


	
	def load_klasses(self, wb):
		out = []
		sheet = wb.sheet_by_name('classi')		
		for nrow in range(1, sheet.nrows):
			row = sheet.row(nrow)
			klass = {}
			klass_field = fix_string(row[0].value)
			klass_value = row[1].value
			klass_description = row[2].value
			klass = { 'field_name' : klass_field, 'value': klass_value, 'description': klass_description}
			out.append(klass)
		
		return out		









class IndicatorManager(object):
	
	def fix_fieldname(self, name):
		"""
		"""
		return unidecode(name).replace(" ", "_")


	def load_meta(self, filename):
		"""
		"""
		wb = load_workbook(filename = filename)
		sheet = wb.get_sheet_by_name('info')
		name = sheet.rows[1][1].value
		code = sheet.rows[0][1].value

		return { 'code':code, 'name':name }


	def load_categories(self):
		"""
		"""


	def load_data(self, filename):
		"""
		"""
		data = []
		#wb = load_workbook(filename = filename)
		wb = openpyxl.reader.excel.load_workbook('test.xlsx')

		sheet_data = wb.get_sheet_by_name('dati')	
		header = sheet_data.rows[0]
		fields = []
		for cell in header:
			fieldname = self.fix_fieldname(cell.value)
			fields.append(fieldname)

		for row in sheet_data.rows[1:]:
			obj = dict()
			for i, cell in enumerate(row):
				field_name = fields[i]
				value  = cell.value
				obj[field_name] = value
			data.append(obj)
		return data




	def import_file(self, filename, update=False):
		"""
		"""
		meta = self.load_meta(filename)
		data = self.load_data(filename)

		if not update:
			self.insert_indicator(meta, data)
		else:
			self.update_indicator(meta, data)


	def insert_indicator(self, meta, data):
		"""
		"""
		ind = IndicatorDescriptor(**meta)
		ind.save()

		for d in data:
			d['code'] = meta['code']
			data_item = IndicatorData(**d)
			data_item.save()
			print data_item, d


	
	def update_indicator(self, meta, data):
		"""
		"""
		ind = IndicatorDescriptor.objects.get(code=meta['code'])
		ind.update(**meta)

	
	def drop_indicator(self, code):
		"""
		"""
		ind = IndicatorDescriptor.objects(code=code)
		ind.delete()

		data_items = IndicatorData.objects(code=code)
		data_items.delete()






